






from django.shortcuts import render,HttpResponse,redirect
from app01 import models
from app01.utils.pagination import Pagination
from app01.utils.form import UserModelForm,MyForm,PrettyModelForm,PrettyEditModelForm
# Create your views here.
def depart_list(request):
    """部门列表"""
    queryset=models.Department.objects.all()

    page_object = Pagination(request, queryset,page_size=2)
    context = {
        'queryset': page_object.page_queryset,  # 分完页的数据
        'page_string': page_object.html()  # 页码
    }


    return render(request,'depart_list.html',context)

def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'depart_add.html')
        # 获取用户提交的数据
    title = request.POST.get("title")
    # 添加到数据库
    models.Department.objects.create(title=title)
    return redirect('/depart/list/')

def depart_delete(request):
    """删除部门"""
    nid=request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect('/depart/list/')

def depart_edit(request,nid):
    """修改部门"""
    if request.method=='GET':
        row_object=models.Department.objects.filter(id=nid).first()
        return render(request,'depart_edit.html',{'row_object':row_object})
    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list/')

from django.core.files.uploadedfile import InMemoryUploadedFile
def depart_multi(request):
    """批量上传（Excel）"""
    from openpyxl import load_workbook

    # 1、获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    print(type(file_object))

    #2、对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    #3、循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text=row[0].value
        #判断是否存在，不存在再添加
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')